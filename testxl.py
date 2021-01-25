import openpyxl
import gol
import re
import string



def checkfile (filename, strcheck):

    if '~' in filename:
        return

    wb = openpyxl.load_workbook(filename)
    # 获取workbook中所有的表格
    sheets = wb.sheetnames

    print(filename)

    # 循环遍历所有sheet
    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        if sheet.title == "Test Steps" or sheet.title == "Test Overview":
            #print('\n\n第' + str(i + 1) + '个sheet: ' + sheet.title + '->>>')
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    #if (string.find(str(sheet.cell(row=r, column=c).value), "PM") )!= -1：
                    ret = re.findall(strcheck,str(sheet.cell(row=r, column=c).value),flags=re.IGNORECASE)
                    if ret:
                        gol.foundstr = "........................   Found Wanted String    .............................."
                        #print("\n find " + strcheck + " in row" + str(r) + " of " + filename)
#                        found_data = [filename,str(r),str(c),ret]
                        found_data = [filename,str(r),str(c),ret,wb[sheets[0]].cell(row=5, column=2).value]
                        gol.a.append(found_data)
                        gol.csv_write.writerow(found_data)
#                        gol.csv_write.writerow([gol.foundstr])
#                    else:
#                        gol.csv_write.writerow([gol.foundstr])




#checkedfile = "t.xlsm"
#needstr = "PM"
#checkfile(checkedfile, needstr)